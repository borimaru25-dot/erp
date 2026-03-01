"""POST /api/files/save-excel-data — Save modified sheet data as Excel."""

import json
import io
import base64
from api._lib.response import json_response, json_error, handle_cors
from api._lib.supabase_client import get_supabase

try:
    import openpyxl
except ImportError:
    openpyxl = None


def handler(request):
    cors = handle_cors(request)
    if cors:
        return cors

    if request.method != "POST":
        return json_error("Method not allowed", 405)

    if openpyxl is None:
        return json_error("openpyxl is not installed", 500)

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, TypeError):
        return json_error("Invalid JSON body")

    filename = body.get("filename", "export.xlsx")
    data = body.get("data")

    if not data or not data.get("headers"):
        return json_error("No data to save")

    try:
        # Create Excel workbook from data
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write headers
        for col_idx, header in enumerate(data["headers"], 1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write rows
        for row_idx, row in enumerate(data["rows"], 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_bytes = buffer.getvalue()

        # Upload to storage
        supabase = get_supabase()
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        storage_path = f"exports/{filename}"
        supabase.storage.from_("files").upload(
            storage_path,
            file_bytes,
            {"content-type": "application/vnd.openxmlformats-officedocument"
                             ".spreadsheetml.sheet",
             "upsert": "true"},
        )

        signed = supabase.storage.from_("files").create_signed_url(
            storage_path, 3600
        )

        return json_response({
            "download_url": signed.get("signedURL", ""),
            "filename": filename,
            "message": "Excel file saved",
        })
    except Exception as e:
        return json_error(f"Save failed: {str(e)}", 500)
